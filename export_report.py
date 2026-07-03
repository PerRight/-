# -*- coding: utf-8 -*-
"""누적 측정 데이터(data/measurements.csv)를 Excel 보고서로 정리한다.

출력: output/report.xlsx — 시트 3개
  1) 회차별 요약    : 회차별 pH·EC 통계, 오염 셀 수, 최고 오염 위치
  2) 오염 셀 목록   : 임계값(pH 6.5~8.5 / EC 500 µS/cm)을 벗어난 셀 상세 (심한 순)
  3) 셀별 전체 데이터: 회차 x 격자 셀별 평균값 전체
오염 셀 행은 빨간 배경으로 강조한다.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from thresholds import exceedance, pollution_reason

DATA_PATH = Path(__file__).parent / "data" / "measurements.csv"
OUT_PATH = Path(__file__).parent / "output" / "report.xlsx"
RED_FILL = PatternFill("solid", start_color="FFF4CCCC")

COL_KR = {
    "run_id": "탐사 회차", "x_m": "X (m)", "y_m": "Y (m)", "depth_m": "수심 (m)",
    "ph_mean": "평균 pH", "ec_mean": "평균 EC (µS/cm)", "n": "측정 횟수", "reason": "판정 사유",
}


def cell_means(df):
    """회차 x (x, y, 수심) 셀별 평균 pH·EC, 측정 수, 오염 판정."""
    g = (
        df.groupby(["run_id", "x_m", "y_m", "depth_m"])
        .agg(ph_mean=("ph", "mean"), ec_mean=("ec", "mean"), n=("ph", "count"))
        .reset_index()
    )
    g["ph_mean"] = g["ph_mean"].round(2)
    g["ec_mean"] = g["ec_mean"].round(0)
    g["reason"] = [pollution_reason(p, e) for p, e in zip(g["ph_mean"], g["ec_mean"])]
    g["severity"] = [
        max(exceedance("ph", p), exceedance("ec", e))
        for p, e in zip(g["ph_mean"], g["ec_mean"])
    ]
    return g


def run_summary(df, cells):
    """회차별 요약 통계 표."""
    rows = []
    for rid, part in df.groupby("run_id"):
        c = cells[cells["run_id"] == rid]
        bad = c[c["reason"] != ""]
        worst = "—"
        if len(bad):
            w = bad.loc[bad["severity"].idxmax()]
            worst = f"(X {w.x_m:.0f}, Y {w.y_m:.0f}, 수심 {w.depth_m} m)"
        rows.append(
            {
                "탐사 회차": int(rid),
                "측정일": str(part["timestamp"].iloc[0])[:10],
                "측정 수": len(part),
                "pH 평균": round(part["ph"].mean(), 2),
                "pH 최소": part["ph"].min(),
                "pH 최대": part["ph"].max(),
                "EC 평균": round(part["ec"].mean(), 0),
                "EC 최소": part["ec"].min(),
                "EC 최대": part["ec"].max(),
                "오염 셀 수": len(bad),
                "최고 오염 위치": worst,
            }
        )
    return pd.DataFrame(rows)


def _fill_red(ws, df, all_rows):
    """오염 행(또는 전체 행)을 빨간 배경으로 강조. 헤더는 1행."""
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        if all_rows or row["판정 사유"] != "":
            for cell in ws[i]:
                cell.fill = RED_FILL


def write_report(df, out_path=OUT_PATH):
    cells = cell_means(df)
    summary = run_summary(df, cells)
    hotspots = (
        cells[cells["reason"] != ""]
        .sort_values("severity", ascending=False)
        .rename(columns=COL_KR)[list(COL_KR.values())]
    )
    detail = cells.rename(columns=COL_KR)[list(COL_KR.values())]
    out_path = Path(out_path)
    out_path.parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        summary.to_excel(xw, sheet_name="회차별 요약", index=False)
        hotspots.to_excel(xw, sheet_name="오염 셀 목록", index=False)
        detail.to_excel(xw, sheet_name="셀별 전체 데이터", index=False)
        _fill_red(xw.book["오염 셀 목록"], hotspots, all_rows=True)
        _fill_red(xw.book["셀별 전체 데이터"], detail, all_rows=False)
    return summary


def main():
    if sys.stdout.encoding and sys.stdout.encoding.lower().replace("-", "") != "utf8":
        sys.stdout.reconfigure(encoding="utf-8")  # Windows 콘솔에서 µ 등 출력 보장
    if not DATA_PATH.exists():
        raise SystemExit("data/measurements.csv 가 없습니다. 먼저 generate_data.py 를 실행하세요.")
    df = pd.read_csv(DATA_PATH)
    if df.empty:
        raise SystemExit("data/measurements.csv 에 데이터가 없습니다. 먼저 generate_data.py 를 실행하세요.")
    summary = write_report(df)
    print(f"저장 완료: {OUT_PATH}")
    print(summary.to_string(index=False))
    print(f"\n총 오염 셀 {int(summary['오염 셀 수'].sum())}개 (회차 {len(summary)}회)")


if __name__ == "__main__":
    main()
