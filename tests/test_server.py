# -*- coding: utf-8 -*-
import dashboard_server as srv


def make_csv(tmp_path):
    p = tmp_path / "m.csv"
    p.write_text(
        "run_id,timestamp,x_m,y_m,depth_m,ph,ec\n"
        "1,2026-06-01 09:00:00,0,0,0.5,7.0,300\n"
        "1,2026-06-01 09:00:20,0,0,0.5,6.0,700\n"
        "2,2026-06-02 09:00:00,10,0,0.5,7.1,310\n",
        encoding="utf-8",
    )
    return p


def test_load_runs(tmp_path):
    runs, meta = srv.load_runs(make_csv(tmp_path))
    assert runs[1][(0.0, 0.0, 0.5)]["ph"] == [13.0, 2]
    assert runs[1][(0.0, 0.0, 0.5)]["ec"] == [1000.0, 2]
    assert meta == [
        {"id": 1, "date": "2026-06-01", "samples": 2},
        {"id": 2, "date": "2026-06-02", "samples": 1},
    ]


def test_load_runs_missing_file(tmp_path):
    runs, meta = srv.load_runs(tmp_path / "none.csv")
    assert runs == {} and meta == []


def test_merge_runs():
    runs = {
        1: {(0.0, 0.0, 0.5): {"ph": [7.0, 1]}},
        2: {(0.0, 0.0, 0.5): {"ph": [6.0, 1]}, (10.0, 0.0, 0.5): {"ec": [300.0, 1]}},
    }
    merged = srv.merge_runs(runs)
    assert merged[(0.0, 0.0, 0.5)]["ph"] == [13.0, 2]
    assert merged[(10.0, 0.0, 0.5)]["ec"] == [300.0, 1]


def test_cells_for_rounds_mean():
    source = {(0.0, 0.0, 0.5): {"ec": [604.0, 2]}}
    assert srv.cells_for(source, "ec", 0) == [[0.0, 0.0, 0.5, 302.0, 2]]


def test_build_summary_counts_either_sensor():
    source = {
        (0.0, 0.0, 0.5): {"ph": [7.0, 1], "ec": [600.0, 1]},   # EC만 초과 → 오염
        (10.0, 0.0, 0.5): {"ph": [7.0, 1], "ec": [300.0, 1]},  # 정상
    }
    s = srv.build_summary(source)
    assert s["polluted_cells"] == 1
    assert s["worst"]["sensor"] == "EC 전기전도도"
    assert s["worst"]["value"] == 600.0
    assert (s["worst"]["x"], s["worst"]["y"], s["worst"]["depth"]) == (0.0, 0.0, 0.5)


def test_build_summary_clean():
    s = srv.build_summary({(0.0, 0.0, 0.5): {"ph": [7.0, 1]}})
    assert s == {"polluted_cells": 0, "worst": None}


def test_build_export_rows():
    source = {(0.0, 0.0, 0.5): {"ph": [6.0, 1], "ec": [600.0, 1]}}
    rows = srv.build_export_rows(source, "1차")
    assert rows == [["1차", 0.0, 0.0, 0.5, 6.0, 600.0, 1, "오염", "pH·EC"]]


def test_build_export_rows_clean():
    source = {(0.0, 0.0, 0.5): {"ph": [7.0, 1], "ec": [300.0, 1]}}
    rows = srv.build_export_rows(source, "실시간")
    assert rows[0][7] == "정상" and rows[0][8] == ""


def test_export_xlsx_roundtrip():
    import io

    from openpyxl import load_workbook

    rows = [["1차", 0.0, 0.0, 0.5, 6.0, 600.0, 1, "오염", "pH·EC"]]
    wb = load_workbook(io.BytesIO(srv.export_xlsx(rows)))
    ws = wb["셀별 데이터"]
    assert ws.max_row == 2
    assert ws.cell(2, 8).value == "오염"


def test_export_csv_has_bom():
    rows = [["1차", 0.0, 0.0, 0.5, 6.0, 600.0, 1, "오염", "pH·EC"]]
    data = srv.export_csv(rows)
    assert data.startswith(b"\xef\xbb\xbf")  # Excel 한글 인식용 BOM
