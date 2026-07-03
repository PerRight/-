# -*- coding: utf-8 -*-
import pandas as pd
import pytest

import export_report
from export_report import cell_means, run_summary, write_report


def make_df():
    return pd.DataFrame(
        {
            "run_id": [1, 1, 1, 2],
            "timestamp": ["2026-06-01 09:00:00"] * 3 + ["2026-06-02 09:00:00"],
            "x_m": [0.0, 0.0, 10.0, 0.0],
            "y_m": [0.0, 0.0, 0.0, 0.0],
            "depth_m": [0.5, 0.5, 0.5, 0.5],
            "ph": [6.0, 6.2, 7.0, 7.1],
            "ec": [600.0, 620.0, 300.0, 310.0],
        }
    )


def test_cell_means_reason():
    cells = cell_means(make_df())
    bad = cells[(cells.run_id == 1) & (cells.x_m == 0.0)].iloc[0]
    assert bad["reason"] == "pH·EC"
    assert bad["n"] == 2
    ok = cells[(cells.run_id == 1) & (cells.x_m == 10.0)].iloc[0]
    assert ok["reason"] == ""


def test_run_summary_counts():
    df = make_df()
    s = run_summary(df, cell_means(df))
    assert list(s["오염 셀 수"]) == [1, 0]
    assert s.iloc[0]["측정일"] == "2026-06-01"
    assert s.iloc[0]["최고 오염 위치"] == "(X 0, Y 0, 수심 0.5 m)"
    assert s.iloc[1]["최고 오염 위치"] == "—"


def test_write_report_sheets(tmp_path):
    from openpyxl import load_workbook

    out = tmp_path / "report.xlsx"
    write_report(make_df(), out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["회차별 요약", "오염 셀 목록", "셀별 전체 데이터"]
    assert wb["오염 셀 목록"].max_row == 2  # 헤더 + 오염 셀 1개
    assert wb["셀별 전체 데이터"].max_row == 4  # 헤더 + 셀 3개

    def is_red(row):
        return all(str(c.fill.start_color.rgb).endswith("F4CCCC") for c in row)

    # 오염 셀 목록: 데이터 행 전부 빨간 배경
    assert is_red(wb["오염 셀 목록"][2])
    # 셀별 전체 데이터: 판정 사유(8열) 오염 행만 빨간 배경
    detail = wb["셀별 전체 데이터"]
    assert detail.cell(row=2, column=8).value == "pH·EC"
    assert is_red(detail[2])
    assert detail.cell(row=3, column=8).value is None  # 정상 셀 (빈 사유)
    assert not is_red(detail[3])


def test_main_empty_csv(tmp_path, monkeypatch):
    empty = tmp_path / "measurements.csv"
    empty.write_text("run_id,timestamp,x_m,y_m,depth_m,ph,ec\n", encoding="utf-8")
    monkeypatch.setattr(export_report, "DATA_PATH", empty)
    with pytest.raises(SystemExit, match="데이터가 없습니다"):
        export_report.main()
